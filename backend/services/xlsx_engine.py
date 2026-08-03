import os
import re
import tempfile
import logging
import copy
import openpyxl
from openpyxl.cell.cell import MergedCell

logger = logging.getLogger(__name__)

_excel_app = None

def get_excel_app():
    global _excel_app
    try:
        import win32com.client
    except (ImportError, ModuleNotFoundError):
        logger.debug("win32com is not available (non-Windows platform).")
        return None

    if _excel_app is not None:
        try:
            _excel_app.Visible = False
            return _excel_app
        except Exception:
            _excel_app = None

    try:
        _excel_app = win32com.client.GetActiveObject("Excel.Application")
    except Exception:
        try:
            _excel_app = win32com.client.Dispatch("Excel.Application")
        except Exception:
            _excel_app = None

    if _excel_app:
        try:
            _excel_app.Visible = False
            _excel_app.DisplayAlerts = False
        except Exception:
            pass
    return _excel_app

def convert_xlsx_to_pdf_local(xlsx_path, pdf_path):
    """
    Converts XLSX to PDF locally using Excel via win32com.
    Reuses a cached Excel instance.
    """
    try:
        import pythoncom
        pythoncom.CoInitialize()
    except (ImportError, ModuleNotFoundError):
        logger.error("pythoncom/win32com not available (likely non-Windows platform).")
        return False
        
    excel = get_excel_app()
    if not excel:
        logger.error("Could not obtain an Excel application instance.")
        return False
        
    try:
        wb = excel.Workbooks.Open(os.path.abspath(xlsx_path))
        wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
        wb.Close(SaveChanges=False)
        logger.info(f"Local Excel conversion successful: {pdf_path}")
        return True
    except Exception as e:
        logger.error(f"Local Excel conversion failed: {e}")
        global _excel_app
        _excel_app = None
        return False

def convert_xlsx_to_pdf_libreoffice(xlsx_path, pdf_path):
    """
    Converts XLSX to PDF using headless LibreOffice (fallback).
    """
    import subprocess
    import shutil
    
    if not shutil.which("libreoffice"):
        logger.debug("LibreOffice is not installed on this system.")
        return False
        
    try:
        outdir = os.path.dirname(pdf_path)
        logger.info(f"Converting {xlsx_path} to PDF via LibreOffice headless...")
        
        cmd = [
            "libreoffice",
            "--headless",
            "-env:UserInstallation=file:///tmp/libreoffice",
            "--convert-to", "pdf",
            "--outdir", outdir,
            xlsx_path
        ]
        
        env = os.environ.copy()
        env["HOME"] = "/tmp"
        
        result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=30, env=env)
        if result.returncode != 0:
            logger.error(f"LibreOffice conversion failed: {result.stderr}")
            return False
            
        default_output_name = os.path.basename(xlsx_path).replace(".xlsx", ".pdf")
        generated_pdf_path = os.path.join(outdir, default_output_name)
        
        if os.path.exists(generated_pdf_path):
            if generated_pdf_path != pdf_path:
                shutil.move(generated_pdf_path, pdf_path)
            logger.info("LibreOffice Excel conversion successful.")
            return True
        return False
    except Exception as e:
        logger.error(f"LibreOffice Excel conversion crashed: {e}")
        return False

def merge_xlsx_template(template_path: str, tokens: dict, output_pdf_path: str = None, output_xlsx_path: str = None, sheet_name: str = None) -> bool:
    """
    Generic, dynamic Excel template engine.
    Supports multi-level loops (Floor/Area/Item), arbitrary template tags, and preserves 100% of template merged cell ranges.
    """
    logger.info(f"Merging XLSX template: {template_path}")
    
    if not os.path.exists(template_path):
        print(f"Error: Excel template missing at {template_path}")
        return False
        
    try:
        wb = openpyxl.load_workbook(template_path)
        matched_sheet = None
        if sheet_name:
            clean_target = str(sheet_name).strip().lower().replace('_', ' ')
            for s_name in wb.sheetnames:
                clean_s = str(s_name).strip().lower().replace('_', ' ')
                if clean_s == clean_target:
                    matched_sheet = s_name
                    break
        
        if matched_sheet:
            ws = wb[matched_sheet]
            for s in list(wb.sheetnames):
                if s != matched_sheet:
                    del wb[s]
        else:
            ws = wb.active
    except Exception as e:
        print(f"Error loading Excel template: {e}")
        return False
    
    floor_header_row = None
    area_header_row = None
    table_header_rows = []
    floor_table_header_rows = []
    area_table_header_rows = []
    item_row = None
    item_summary_row = None
    area_footer_row = None
    floor_footer_row = None
    
    spacer_row = None
    
    # 1. Scan Column A (and cell values) for control tags
    for r in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=r, column=1)
        cell_a_val = str(cell_a.value or '').strip()
        
        # Check merged cell top-left value if current cell value is empty
        if not cell_a_val and type(cell_a).__name__ == 'MergedCell':
            for m in ws.merged_cells.ranges:
                if m.min_row <= r <= m.max_row and m.min_col <= 1 <= m.max_col:
                    top_left = ws.cell(row=m.min_row, column=m.min_col)
                    cell_a_val = str(top_left.value or '').strip()
                    break
        
        if "[FLOOR_HEADER]" in cell_a_val or "{{#floor}}" in cell_a_val:
            floor_header_row = r
        elif "[AREA_HEADER]" in cell_a_val or "{{#area}}" in cell_a_val:
            area_header_row = r
        elif "[FLOOR_TABLE_HEADER]" in cell_a_val:
            floor_table_header_rows.append(r)
        elif "[AREA_TABLE_HEADER]" in cell_a_val:
            area_table_header_rows.append(r)
        elif "[TABLE_HEADER]" in cell_a_val:
            table_header_rows.append(r)
        elif "[SPACER]" in cell_a_val or "[SPACE]" in cell_a_val:
            spacer_row = r
        elif "[ITEM_SUMMARY]" in cell_a_val:
            if item_summary_row is None:
                item_summary_row = r
        elif "[ITEM]" in cell_a_val or "{{item." in cell_a_val:
            if item_row is None:
                item_row = r
        elif "[AREA_FOOTER]" in cell_a_val or "{{/area}}" in cell_a_val:
            area_footer_row = r
        elif "[FLOOR_FOOTER]" in cell_a_val or "{{/floor}}" in cell_a_val:
            floor_footer_row = r

    has_tagged_loop = (floor_header_row is not None or area_header_row is not None or item_row is not None or item_summary_row is not None or area_footer_row is not None or floor_footer_row is not None or len(table_header_rows) > 0 or len(floor_table_header_rows) > 0 or len(area_table_header_rows) > 0 or spacer_row is not None)
    
    if has_tagged_loop:
        control_rows = list(filter(None, [floor_header_row, area_header_row, item_row, item_summary_row, area_footer_row, floor_footer_row, spacer_row] + table_header_rows + floor_table_header_rows + area_table_header_rows))
        min_ctrl_row = min(control_rows)
        max_ctrl_row = max(control_rows)
        original_ctrl_height = max_ctrl_row - min_ctrl_row + 1

        # Snapshot original row heights across the whole sheet before doing any operations
        original_row_heights = {r: dim.height for r, dim in ws.row_dimensions.items() if dim.height is not None}

        # Snapshot all template merged cell ranges strictly outside control rows before doing any row operations
        fixed_merges_above = []
        fixed_merges_below = []
        for m in list(ws.merged_cells.ranges):
            if m.max_row < min_ctrl_row:
                fixed_merges_above.append((m.min_row, m.max_row, m.min_col, m.max_col))
            elif m.min_row > max_ctrl_row:
                fixed_merges_below.append((m.min_row, m.max_row, m.min_col, m.max_col))

        def get_row_design(row_num):
            if row_num is None: return []
            cells = []
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=c)
                cells.append({
                    "col": c,
                    "value": cell.value,
                    "number_format": cell.number_format,
                    "font": copy.copy(cell.font) if cell.font else None,
                    "fill": copy.copy(cell.fill) if cell.fill else None,
                    "border": copy.copy(cell.border) if cell.border else None,
                    "alignment": copy.copy(cell.alignment) if cell.alignment else None
                })
            return cells

        floor_header_design = get_row_design(floor_header_row)
        area_header_design = get_row_design(area_header_row)
        table_header_designs = [get_row_design(r) for r in table_header_rows]
        floor_table_header_designs = [get_row_design(r) for r in floor_table_header_rows]
        area_table_header_designs = [get_row_design(r) for r in area_table_header_rows]
        spacer_design = get_row_design(spacer_row) if spacer_row else []
        item_design = get_row_design(item_row) if item_row else []
        item_summary_design = get_row_design(item_summary_row) if item_summary_row else []
        area_footer_design = get_row_design(area_footer_row) if area_footer_row else []
        floor_footer_design = get_row_design(floor_footer_row) if floor_footer_row else []

        # Capture control row merges from original template before unmerging
        def get_row_merges(row_num):
            if row_num is None: return []
            merges = []
            for m in list(ws.merged_cells.ranges):
                if m.min_row == row_num:
                    merges.append((m.min_col, m.max_col))
            return merges

        floor_header_merges = get_row_merges(floor_header_row)
        area_header_merges = get_row_merges(area_header_row)
        spacer_merges = get_row_merges(spacer_row) if spacer_row else []
        item_merges = []
        item_summary_merges = get_row_merges(item_summary_row) if item_summary_row else []
        area_footer_merges = get_row_merges(area_footer_row)
        floor_footer_merges = get_row_merges(floor_footer_row)

        # Unmerge all control block ranges and specifically unmerge item_row to prevent insert_rows from inheriting merges
        for m in list(ws.merged_cells.ranges):
            if (item_row and m.min_row <= item_row <= m.max_row) or (m.min_row >= min_ctrl_row and m.max_row <= max_ctrl_row):
                try: ws.unmerge_cells(str(m))
                except Exception: pass

        row_heights = {}
        for r_idx, r_num in [("floor_header", floor_header_row), ("area_header", area_header_row), ("spacer", spacer_row), ("item", item_row), ("item_summary", item_summary_row), ("area_footer", area_footer_row), ("floor_footer", floor_footer_row)]:
            if r_num and original_row_heights.get(r_num):
                row_heights[r_idx] = original_row_heights.get(r_num)
                
        table_header_heights = [original_row_heights.get(r) for r in table_header_rows if original_row_heights.get(r)]
        floor_table_header_heights = [original_row_heights.get(r) for r in floor_table_header_rows if original_row_heights.get(r)]
        area_table_header_heights = [original_row_heights.get(r) for r in area_table_header_rows if original_row_heights.get(r)]

        # Delete placeholder block rows (min_ctrl_row to max_ctrl_row)
        ws.delete_rows(min_ctrl_row, original_ctrl_height)

        def apply_design(target_row, design_list, value_replacements=None, row_height=None):
            ws.insert_rows(target_row, 1)
            if row_height:
                ws.row_dimensions[target_row].height = row_height

            for cell_def in design_list:
                col_idx = cell_def["col"]
                target_cell = ws.cell(row=target_row, column=col_idx)
                
                if cell_def["font"]: target_cell.font = copy.copy(cell_def["font"])
                if cell_def["fill"]: target_cell.fill = copy.copy(cell_def["fill"])
                if cell_def["border"]: target_cell.border = copy.copy(cell_def["border"])
                if cell_def["alignment"]: target_cell.alignment = copy.copy(cell_def["alignment"])
                if cell_def["number_format"]: target_cell.number_format = cell_def["number_format"]
                
                val = cell_def["value"]
                if value_replacements and val:
                    val_str = str(val)
                    # Strip block tags if present inside cell text
                    val_str = val_str.replace("{{#area}}", "").replace("{{/area}}", "").replace("{{#floor}}", "").replace("{{/floor}}", "")
                    sorted_repls = sorted(value_replacements.items(), key=lambda x: len(x[0]), reverse=True)
                    for k, v in sorted_repls:
                        val_str = val_str.replace(k, str(v if v is not None else ''))
                    
                    try:
                        stripped_val = val_str.strip()
                        if stripped_val.startswith('R '):
                            clean_val = stripped_val.replace('R ', '').replace(',', '').strip()
                            if re.match(r'^-?\d+(?:\.\d+)?$', clean_val):
                                target_cell.value = float(clean_val)
                            else:
                                target_cell.value = val_str
                        elif re.match(r'^-?\d+(?:\.\d+)?$', stripped_val):
                            if '.' in stripped_val:
                                target_cell.value = float(stripped_val)
                            else:
                                target_cell.value = int(stripped_val)
                        else:
                            target_cell.value = val_str
                    except Exception:
                        target_cell.value = val_str
                else:
                    if val and isinstance(val, str):
                        val_cleaned = val.replace("{{#area}}", "").replace("{{/area}}", "").replace("{{#floor}}", "").replace("{{/floor}}", "")
                        target_cell.value = val_cleaned
                    else:
                        target_cell.value = val
                    
            return target_row

        def clean_price(val_in):
            if not val_in: return 0.0
            if isinstance(val_in, (int, float)): return float(val_in)
            val_s = str(val_in).replace("R", "").replace(",", "").strip()
            try: return float(val_s)
            except ValueError: return 0.0

        def summarize_item_list(raw_items):
            grouped = {}
            for it in raw_items:
                if it.get("isSpacer") or it.get("type") == "SPACER": continue
                one_code = str(it.get("oneOneCode") or it.get("oneonecode") or it.get("one_one_code") or '').strip()
                item_type = str(it.get("type") or it.get("planCode") or it.get("plan_code") or '').strip()
                key = (one_code.lower(), item_type.lower())
                qty = float(it.get("qty") or 0)
                unit_cost = clean_price(it.get("unitCost"))
                unit_retail = clean_price(it.get("unitRetail"))
                
                if key not in grouped:
                    grouped[key] = {
                        **copy.deepcopy(it),
                        "qty": qty,
                        "unitCost": unit_cost,
                        "unitRetail": unit_retail,
                        "totalRetail": clean_price(it.get("totalRetail"))
                    }
                else:
                    grouped[key]["qty"] += qty
                    grouped[key]["totalRetail"] += clean_price(it.get("totalRetail"))
            
            # Format summarized items for output
            result = []
            for g_item in grouped.values():
                result.append({
                    **g_item,
                    "qty": int(g_item["qty"]) if g_item["qty"].is_integer() else round(g_item["qty"], 2),
                    "totalRetail": f"R {g_item['totalRetail']:,.2f}"
                })
            return result

        curr_row = min_ctrl_row
        inserted_rows_count = 0

        # Helper to render items or item summary
        def render_item_rows(items_to_render):
            nonlocal curr_row, inserted_rows_count
            if item_design:
                for idx, item in enumerate(items_to_render):
                    if item.get("isSpacer") or item.get("type") == "SPACER":
                        if spacer_design:
                            apply_design(curr_row, spacer_design, {}, row_heights.get("spacer") or 9.0)
                            for min_c, max_c in spacer_merges:
                                try: ws.merge_cells(start_row=curr_row, start_column=min_c, end_row=curr_row, end_column=max_c)
                                except Exception: pass
                        elif item_design:
                            # Apply thin space row using item_design styling across all table columns
                            ws.insert_rows(curr_row, 1)
                            ws.row_dimensions[curr_row].height = 9.0
                            for cell_def in item_design:
                                col_idx = cell_def["col"]
                                target_cell = ws.cell(row=curr_row, column=col_idx)
                                if cell_def["fill"]: target_cell.fill = copy.copy(cell_def["fill"])
                                if cell_def["border"]: target_cell.border = copy.copy(cell_def["border"])
                                target_cell.value = ""
                        else:
                            ws.insert_rows(curr_row, 1)
                            ws.row_dimensions[curr_row].height = 9.0
                        curr_row += 1
                        inserted_rows_count += 1
                        continue

                    repls = {"{{index}}": str(idx + 1)}
                    for k, v in item.items():
                        str_val = str(v if v is not None else '')
                        repls["{{item." + str(k) + "}}"] = str_val
                        repls["{{" + str(k) + "}}"] = str_val
                        k_lower = str(k).lower()
                        repls["{{item." + k_lower + "}}"] = str_val
                        repls["{{" + k_lower + "}}"] = str_val

                    one_code = item.get("oneOneCode") or ''
                    item_type = item.get("type") or item.get("planCode") or item.get("plan_code") or ''
                    repls["{{item.oneOneCode}}"] = one_code
                    repls["{{item.oneonecode}}"] = one_code
                    repls["{{item.one_one_code}}"] = one_code
                    repls["{{oneOneCode}}"] = one_code
                    repls["{{oneonecode}}"] = one_code
                    repls["{{one_one_code}}"] = one_code
                    repls["{{item.type}}"] = item_type
                    repls["{{item.planCode}}"] = item_type
                    repls["{{item.plan_code}}"] = item_type
                    repls["{{type}}"] = item_type
                    repls["{{planCode}}"] = item_type
                    repls["{{plan_code}}"] = item_type
                        
                    apply_design(curr_row, item_design, repls, row_heights.get("item"))
                    curr_row += 1
                    inserted_rows_count += 1
            elif item_summary_design:
                summed_items = summarize_item_list(items_to_render)
                for idx, item in enumerate(summed_items):
                    repls = {"{{index}}": str(idx + 1)}
                    for k, v in item.items():
                        str_val = str(v if v is not None else '')
                        repls["{{item." + str(k) + "}}"] = str_val
                        repls["{{" + str(k) + "}}"] = str_val
                        k_lower = str(k).lower()
                        repls["{{item." + k_lower + "}}"] = str_val
                        repls["{{" + k_lower + "}}"] = str_val

                    one_code = item.get("oneOneCode") or ''
                    item_type = item.get("type") or item.get("planCode") or item.get("plan_code") or ''
                    repls["{{item.oneOneCode}}"] = one_code
                    repls["{{item.oneonecode}}"] = one_code
                    repls["{{item.one_one_code}}"] = one_code
                    repls["{{oneOneCode}}"] = one_code
                    repls["{{oneonecode}}"] = one_code
                    repls["{{one_one_code}}"] = one_code
                    repls["{{item.type}}"] = item_type
                    repls["{{item.planCode}}"] = item_type
                    repls["{{item.plan_code}}"] = item_type
                    repls["{{type}}"] = item_type
                    repls["{{planCode}}"] = item_type
                    repls["{{plan_code}}"] = item_type
                        
                    apply_design(curr_row, item_summary_design, repls, row_heights.get("item_summary"))
                    for min_c, max_c in item_summary_merges:
                        try: ws.merge_cells(start_row=curr_row, start_column=min_c, end_row=curr_row, end_column=max_c)
                        except Exception: pass
                    curr_row += 1
                    inserted_rows_count += 1

        # Check if sheet uses Section Headers or Section Footers ([FLOOR_HEADER], [AREA_HEADER], [AREA_FOOTER], [FLOOR_FOOTER])
        has_section_headers = (floor_header_design or area_header_design or area_footer_design or floor_footer_design)

        if has_section_headers:
            floors = tokens.get("floors", [])
            for f in floors:
                floor_name = str(f.get("name", "")).strip()
                valid_areas = [a for a in f.get("areas", []) if len(a.get("items", [])) > 0]
                if not valid_areas: continue

                # 1. Insert Floor Header
                if floor_header_design:
                    apply_design(curr_row, floor_header_design, {"{{floor.name}}": floor_name}, row_heights.get("floor_header"))
                    for min_c, max_c in floor_header_merges:
                        try: ws.merge_cells(start_row=curr_row, start_column=min_c, end_row=curr_row, end_column=max_c)
                        except Exception: pass
                    curr_row += 1
                    inserted_rows_count += 1

                # 1b. Insert Floor Table Header (Repeated ONCE per Floor)
                eff_floor_headers = floor_table_header_designs if floor_table_header_designs else (table_header_designs if (floor_header_design and not area_header_design) else [])
                eff_floor_heights = floor_table_header_heights if floor_table_header_heights else table_header_heights
                for h_idx, h_design in enumerate(eff_floor_headers):
                    h_height = eff_floor_heights[h_idx] if h_idx < len(eff_floor_heights) else None
                    apply_design(curr_row, h_design, None, h_height)
                    curr_row += 1
                    inserted_rows_count += 1

                for a in valid_areas:
                    area_name = str(a.get("name", "")).strip()
                    # 2. Insert Area Header
                    if area_header_design:
                        apply_design(curr_row, area_header_design, {"{{area.name}}": area_name}, row_heights.get("area_header"))
                        for min_c, max_c in area_header_merges:
                            try: ws.merge_cells(start_row=curr_row, start_column=min_c, end_row=curr_row, end_column=max_c)
                            except Exception: pass
                        curr_row += 1
                        inserted_rows_count += 1

                    # 3. Insert Area Table Headers (Repeated per Area)
                    eff_area_headers = area_table_header_designs if area_table_header_designs else (table_header_designs if area_header_design else [])
                    eff_area_heights = area_table_header_heights if area_table_header_heights else table_header_heights
                    for h_idx, h_design in enumerate(eff_area_headers):
                        h_height = eff_area_heights[h_idx] if h_idx < len(eff_area_heights) else None
                        apply_design(curr_row, h_design, None, h_height)
                        curr_row += 1
                        inserted_rows_count += 1

                    # 4. Insert Items / Item Summary inside Area
                    render_item_rows(a.get("items", []))

                    # 5. Insert Area Footer / Subtotal
                    if area_footer_design:
                        area_subtotal = sum(clean_price(item.get("totalRetail")) for item in a.get("items", []))
                        subtotal_repls = {
                            "{{SUBTOTAL}}": f"R {area_subtotal:,.2f}",
                            "{{area.name}}": area_name,
                            "{{floor.name}}": floor_name
                        }
                        apply_design(curr_row, area_footer_design, subtotal_repls, row_heights.get("area_footer"))
                        
                        for cell_def in area_footer_design:
                            cell_val_s = str(cell_def["value"] or '')
                            tc = ws.cell(row=curr_row, column=cell_def["col"])
                            if cell_val_s.strip() == "{{SUBTOTAL}}":
                                tc.value = area_subtotal
                                if cell_def["number_format"] and cell_def["number_format"] != 'General':
                                    tc.number_format = cell_def["number_format"]
                                else:
                                    tc.number_format = '"R"#,##0.00'
                            elif "{{" in cell_val_s:
                                val_t = cell_val_s
                                for k, v in subtotal_repls.items():
                                    val_t = val_t.replace(k, str(v))
                                tc.value = val_t
                                
                        for min_c, max_c in area_footer_merges:
                            try: ws.merge_cells(start_row=curr_row, start_column=min_c, end_row=curr_row, end_column=max_c)
                            except Exception: pass

                        curr_row += 1
                        inserted_rows_count += 1

                # 6. Insert Floor Footer
                if floor_footer_design:
                    floor_subtotal = sum(clean_price(item.get("totalRetail")) for area in valid_areas for item in area.get("items", []))
                    floor_subtotal_repls = {
                        "{{SUBTOTAL}}": f"R {floor_subtotal:,.2f}",
                        "{{floor.name}}": floor_name
                    }
                    apply_design(curr_row, floor_footer_design, floor_subtotal_repls, row_heights.get("floor_footer"))
                    
                    for cell_def in floor_footer_design:
                        cell_val_s = str(cell_def["value"] or '')
                        tc = ws.cell(row=curr_row, column=cell_def["col"])
                        if cell_val_s.strip() == "{{SUBTOTAL}}":
                            tc.value = floor_subtotal
                            if cell_def["number_format"] and cell_def["number_format"] != 'General':
                                tc.number_format = cell_def["number_format"]
                            else:
                                tc.number_format = '"R"#,##0.00'
                        elif "{{" in cell_val_s:
                            val_t = cell_val_s
                            for k, v in floor_subtotal_repls.items():
                                val_t = val_t.replace(k, str(v))
                            tc.value = val_t

                    for min_c, max_c in floor_footer_merges:
                        try: ws.merge_cells(start_row=curr_row, start_column=min_c, end_row=curr_row, end_column=max_c)
                        except Exception: pass
                    curr_row += 1
                    inserted_rows_count += 1

        else:
            # Flat Table Loop (No [FLOOR_HEADER] or [AREA_HEADER])
            # 1. Render Table Header ONCE (Not repeated)
            for h_idx, h_design in enumerate(table_header_designs):
                h_height = table_header_heights[h_idx] if h_idx < len(table_header_heights) else None
                apply_design(curr_row, h_design, None, h_height)
                curr_row += 1
                inserted_rows_count += 1

            # 2. Render all items across the entire order
            all_order_items = tokens.get("items", [])
            render_item_rows(all_order_items)

        delta_rows = inserted_rows_count - original_ctrl_height

        # Clear and shift row heights for fixed rows below min_ctrl_row
        for r in list(ws.row_dimensions.keys()):
            if r >= min_ctrl_row:
                ws.row_dimensions[r].height = None

        for r, h in original_row_heights.items():
            if r < min_ctrl_row:
                ws.row_dimensions[r].height = h
            elif r > max_ctrl_row:
                ws.row_dimensions[r + delta_rows].height = h

        # Re-apply all fixed merged cell ranges above the loop
        for min_r, max_r, min_c, max_c in fixed_merges_above:
            try: ws.merge_cells(start_row=min_r, start_column=min_c, end_row=max_r, end_column=max_c)
            except Exception: pass

        # Re-apply all fixed merged cell ranges below the loop with exact delta_rows offset
        for min_r, max_r, min_c, max_c in fixed_merges_below:
            try: ws.merge_cells(start_row=min_r + delta_rows, start_column=min_c, end_row=max_r + delta_rows, end_column=max_c)
            except Exception: pass
            
    else:
        # Standard Flat Loop Fallback Logic
        items = tokens.get("items", [])
        loop_start_row = None
        loop_end_row = None
        
        for r in range(1, ws.max_row + 1):
            cell_a_val = str(ws.cell(row=r, column=1).value or '').strip()
            if cell_a_val == "[ITEM]":
                loop_start_row = r
                loop_end_row = r
                break
                
        if loop_start_row is None:
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r, column=c)
                    if isinstance(cell, MergedCell): continue
                    val = str(cell.value or '')
                    if "{{#each items}}" in val:
                        loop_start_row = r
                    if "{{/each}}" in val:
                        loop_end_row = r
                        
        if loop_start_row is not None and loop_end_row is not None:
            for m in list(ws.merged_cells.ranges):
                if m.min_row >= loop_start_row and m.max_row <= loop_end_row:
                    try: ws.unmerge_cells(str(m))
                    except Exception: pass

            template_item_row = loop_start_row + 1 if loop_start_row != loop_end_row else loop_start_row
            item_row_height = ws.row_dimensions[template_item_row].height

            item_design = []
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=template_item_row, column=c)
                item_design.append({
                    "col": c,
                    "value": cell.value,
                    "number_format": cell.number_format,
                    "font": copy.copy(cell.font) if cell.font else None,
                    "fill": copy.copy(cell.fill) if cell.fill else None,
                    "border": copy.copy(cell.border) if cell.border else None,
                    "alignment": copy.copy(cell.alignment) if cell.alignment else None
                })

            for r in range(loop_end_row, loop_start_row - 1, -1):
                try: ws.delete_rows(r, 1)
                except Exception: pass

            curr_row = loop_start_row
            for idx, item in enumerate(items):
                ws.insert_rows(curr_row, 1)
                if item_row_height:
                    ws.row_dimensions[curr_row].height = item_row_height

                for cell_def in item_design:
                    target_cell = ws.cell(row=curr_row, column=cell_def["col"])
                    if isinstance(target_cell, MergedCell): continue

                    if cell_def["font"]: target_cell.font = copy.copy(cell_def["font"])
                    if cell_def["fill"]: target_cell.fill = copy.copy(cell_def["fill"])
                    if cell_def["border"]: target_cell.border = copy.copy(cell_def["border"])
                    if cell_def["alignment"]: target_cell.alignment = copy.copy(cell_def["alignment"])
                    if cell_def["number_format"]: target_cell.number_format = cell_def["number_format"]

                    val_str = str(cell_def["value"] or '')
                    if val_str:
                        item_repls = {"{{index}}": str(idx + 1)}
                        for k, v in item.items():
                            str_val = str(v if v is not None else '')
                            item_repls["{{item." + str(k) + "}}"] = str_val
                            item_repls["{{" + str(k) + "}}"] = str_val
                            
                            k_lower = str(k).lower()
                            item_repls["{{item." + k_lower + "}}"] = str_val
                            item_repls["{{" + k_lower + "}}"] = str_val

                        one_code = item.get("oneOneCode") or ''
                        item_type = item.get("type") or item.get("planCode") or item.get("plan_code") or ''

                        item_repls["{{item.oneOneCode}}"] = one_code
                        item_repls["{{item.oneonecode}}"] = one_code
                        item_repls["{{item.one_one_code}}"] = one_code
                        item_repls["{{oneOneCode}}"] = one_code
                        item_repls["{{oneonecode}}"] = one_code
                        item_repls["{{one_one_code}}"] = one_code

                        item_repls["{{item.type}}"] = item_type
                        item_repls["{{item.planCode}}"] = item_type
                        item_repls["{{item.plan_code}}"] = item_type
                        item_repls["{{type}}"] = item_type
                        item_repls["{{planCode}}"] = item_type
                        item_repls["{{plan_code}}"] = item_type

                        sorted_item_repls = sorted(item_repls.items(), key=lambda x: len(x[0]), reverse=True)
                        for k, v in sorted_item_repls:
                            val_str = val_str.replace(k, v)

                        try:
                            stripped_val = val_str.strip()
                            if stripped_val.startswith('R '):
                                clean_val = stripped_val.replace('R ', '').replace(',', '').strip()
                                if re.match(r'^-?\d+(?:\.\d+)?$', clean_val):
                                    target_cell.value = float(clean_val)
                                else:
                                    target_cell.value = val_str
                            elif re.match(r'^-?\d+(?:\.\d+)?$', stripped_val):
                                if '.' in stripped_val:
                                    target_cell.value = float(stripped_val)
                                else:
                                    target_cell.value = int(stripped_val)
                            else:
                                target_cell.value = val_str
                        except Exception:
                            target_cell.value = val_str
                    else:
                        target_cell.value = None
                curr_row += 1
    
    # 2. Second Pass: Find and replace single global variables in other rows (Col 1 to Max Column)
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if type(cell).__name__ == 'MergedCell': continue
            val = str(cell.value or '')
            if val and ("{?" in val or "{{" in val):
                val = val.replace("{{#floor}}", "").replace("{{#area}}", "").replace("{{/area}}", "").replace("{{/floor}}", "")
                
                for k, v in tokens.items():
                    if not isinstance(v, (list, dict)):
                        val = val.replace("{{" + str(k) + "}}", str(v if v is not None else ''))
                        val = val.replace("{?" + str(k) + "?}", str(v if v is not None else ''))
                
                val = re.sub(r'\{\{[^}]+\}\}', '', val)
                
                if val != "":
                    try:
                        stripped_val = val.strip()
                        if stripped_val.startswith('R '):
                            clean_val = stripped_val.replace('R ', '').replace('R', '').replace(',', '').strip()
                            if re.match(r'^-?\d+(?:\.\d+)?$', clean_val):
                                cell.value = float(clean_val)
                                cell.number_format = '"R"#,##0.00'
                            else:
                                cell.value = val
                        elif re.match(r'^-?\d+(?:\.\d+)?$', stripped_val):
                            if '.' in stripped_val:
                                cell.value = float(stripped_val)
                                cell.number_format = '"R"#,##0.00'
                            else:
                                cell.value = int(stripped_val)
                        else:
                            cell.value = val
                    except Exception:
                        cell.value = val
                else:
                    pass

    # Calculate exact max occupied column letter dynamically
    max_col_idx = 2
    for r in range(1, ws.max_row + 1):
        for c in range(ws.max_column, 0, -1):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None and str(cell.value).strip() != '':
                if c > max_col_idx:
                    max_col_idx = c
                break

    from openpyxl.utils import get_column_letter
    max_col_letter = get_column_letter(max_col_idx)

    # 3. Third Pass: Clear Column A text (control markers)
    has_column_a_control_tags = False
    for r in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=r, column=1)
        if type(cell_a).__name__ != 'MergedCell':
            val_a = str(cell_a.value or '').strip()
            if "[" in val_a and "]" in val_a:
                has_column_a_control_tags = True
                cell_a.value = None

    if has_column_a_control_tags:
        if not getattr(ws, '_images', None):
            ws.column_dimensions['A'].hidden = True
        start_col = "B"
    else:
        start_col = "A"

    ws.print_area = f"{start_col}1:{max_col_letter}{ws.max_row}"

    # 4. Fit-To-Page setup for PDF conversion
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    if ws.sheet_properties and ws.sheet_properties.pageSetUpPr:
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    if output_xlsx_path:
        wb.save(output_xlsx_path)
        return True

    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    temp_xlsx_path = temp_xlsx.name
    temp_xlsx.close()
    
    wb.save(temp_xlsx_path)
    
    success = convert_xlsx_to_pdf_local(temp_xlsx_path, output_pdf_path)
    if not success:
        success = convert_xlsx_to_pdf_libreoffice(temp_xlsx_path, output_pdf_path)
        
    try:
        os.remove(temp_xlsx_path)
    except Exception:
        pass
        
    return success
