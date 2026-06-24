import sys
import os
import re
import tempfile
import requests
import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

# Add backend directory to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.cloud_sql import Base, clean_database_url
from models.orm_models import Project, Client, BOQ, BOQItem, ProjectFolder
import models.orm_models

SPREADSHEET_ID = "1N2lDgBecJUs1SMwfGbjRqBC8eeyN1QLFbXRWRUTHjeU"
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx"

def clean_number(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    val_str = str(val).strip()
    cleaned = re.sub(r"[^\d\.]", "", val_str.replace(",", "."))
    if cleaned.count(".") > 1:
        parts = cleaned.split(".")
        cleaned = f"{parts[0]}.{parts[1]}"
    try:
        return float(cleaned) if cleaned else 0.0
    except ValueError:
        return 0.0

def clean_int(val):
    if val is None:
        return 0
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        return int(val)
    val_str = str(val).strip()
    cleaned = re.sub(r"[^\d]", "", val_str)
    try:
        return int(cleaned) if cleaned else 0
    except ValueError:
        return 0

def find_value_by_label(sheet, label):
    # Search first 15 rows and first 12 columns
    for r in range(1, 15):
        for c in range(1, 12):
            try:
                val = sheet.cell(row=r, column=c).value
                if val and label.strip().lower() in str(val).strip().lower():
                    # Return the value in the next column
                    return sheet.cell(row=r, column=c+1).value
            except Exception:
                pass
    return None

def run():
    target_url = os.getenv("DATABASE_URL")
    if not target_url:
        print("ERROR: DATABASE_URL is not set.")
        return

    target_url = clean_database_url(target_url)
    engine = create_engine(target_url, pool_pre_ping=True)
    SessionLocal = sessionmaker(bind=engine)
    db = None

    # 1. Download spreadsheet as Excel binary
    print("Downloading Google Sheet as Excel binary (anonymous web access)...")
    try:
        response = requests.get(EXPORT_URL, timeout=120)
        response.raise_for_status()
    except Exception as e:
        print(f"Error downloading spreadsheet: {e}")
        return

    # Save to temp file
    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_file.write(response.content)
    tmp_file.close()

    print("Loading workbook (read-only mode)...")
    try:
        wb = openpyxl.load_workbook(tmp_file.name, data_only=True, read_only=True)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        os.unlink(tmp_file.name)
        return

    # Find the index of the 'Template' tab
    sheet_names = wb.sheetnames
    template_idx = -1
    for idx, name in enumerate(sheet_names):
        if name.strip().lower() == 'template':
            template_idx = idx
            break

    if template_idx == -1:
        print("Error: Could not find 'Template' tab in the spreadsheet.")
        os.unlink(tmp_file.name)
        return

    target_sheets = sheet_names[template_idx + 1:]
    print(f"Found {len(target_sheets)} order tabs to import after 'Template'.")

    for title in target_sheets:
        db = SessionLocal()
        try:
            print(f"\nProcessing tab: '{title}'...")
            sheet = wb[title]

            # Extract cells dynamically by matching labels
            company_name = find_value_by_label(sheet, "Company:")
            project_name = find_value_by_label(sheet, "Project:")
            sales_rep = find_value_by_label(sheet, "Sale Rep Name:") or find_value_by_label(sheet, "Sale Rep:")
            date_created = find_value_by_label(sheet, "Date Created:")
            pf_number = find_value_by_label(sheet, "PF Number & Date:") or find_value_by_label(sheet, "PF Number")

            # Clean string values
            company_name = str(company_name).strip() if company_name else "Steyn City"
            project_name = str(project_name).strip() if project_name else title
            sales_rep = str(sales_rep).strip() if sales_rep else "Dani"
            date_created = str(date_created).strip() if date_created else ""

            print(f"  -> Project Name: {project_name}")
            print(f"  -> Company Name: {company_name}")
            print(f"  -> Sales Rep: {sales_rep}")
            print(f"  -> Date Created: {date_created}")

            # 1. Create or Find Client
            client = db.query(Client).filter(Client.company == company_name).first()
            if not client:
                client = Client(
                    name=company_name,
                    company=company_name,
                    status="Active"
                )
                db.add(client)
                db.commit()
                db.refresh(client)

            # 2. Create Project
            project = db.query(Project).filter(Project.name == project_name).first()
            if not project:
                project = Project(
                    name=project_name,
                    client_id=client.id,
                    status="On track",
                    complete_status="Ongoing"
                )
                db.add(project)
                db.commit()
                db.refresh(project)
                
                # Setup default folders for the project
                fld_design = ProjectFolder(
                    project_id=project.id,
                    gdrive_folder_id=f"gdrive-fld-design-{project.id}",
                    parent_id=None,
                    sort_order=1,
                    name="Stage 2: Design Files"
                )
                fld_supply = ProjectFolder(
                    project_id=project.id,
                    gdrive_folder_id=f"gdrive-fld-supply-{project.id}",
                    parent_id=None,
                    sort_order=2,
                    name="Stage 3: Product Supply"
                )
                db.add(fld_design)
                db.add(fld_supply)
                db.commit()
                db.refresh(fld_design)
                
                fld_cad = ProjectFolder(
                    project_id=project.id,
                    gdrive_folder_id=f"gdrive-fld-cad-{project.id}",
                    parent_id=fld_design.id,
                    sort_order=1,
                    name="CAD Layouts"
                )
                db.add(fld_cad)
                db.commit()

            # 3. Create BOQ
            # Clear old BOQs
            old_boqs = db.query(BOQ).filter(BOQ.project_id == project.id).all()
            for ob in old_boqs:
                db.query(BOQItem).filter(BOQItem.boq_id == ob.id).delete()
                db.delete(ob)
            db.commit()

            boq = BOQ(
                project_id=project.id,
                status="approved",
                created_at=date_created
            )
            db.add(boq)
            db.commit()
            db.refresh(boq)

            # 4. Find table headers dynamically
            header_row = -1
            col_indices = {}
            for r in range(8, 15):
                try:
                    row_vals = [str(sheet.cell(row=r, column=c).value or "").strip().lower() for c in range(1, 15)]
                    if any("qty" in val for val in row_vals):
                        header_row = r
                        for c in range(1, 15):
                            val = str(sheet.cell(row=r, column=c).value or "").strip().lower()
                            if "qty" in val:
                                col_indices["qty"] = c
                            elif "code" in val and not "model" in val:
                                col_indices["code"] = c
                            elif "description" in val:
                                col_indices["desc"] = c
                            elif "price" in val or "unit price" in val:
                                col_indices["price"] = c
                            elif "cost" in val:
                                col_indices["cost"] = c
                        break
                except Exception:
                    pass

            if header_row == -1 or "qty" not in col_indices or "code" not in col_indices:
                print("  -> Warning: Could not find table headers. Skipping line items.")
                continue

            # 5. Parse Line Items starting from header_row + 1
            items_count = 0
            for row in range(header_row + 1, 150):
                try:
                    qty_val = sheet.cell(row=row, column=col_indices["qty"]).value
                    code_val = sheet.cell(row=row, column=col_indices["code"]).value
                    desc_val = sheet.cell(row=row, column=col_indices.get("desc", 4)).value

                    # Skip if we encounter datetimes directly
                    import datetime
                    if isinstance(qty_val, (datetime.datetime, datetime.date)) or isinstance(code_val, (datetime.datetime, datetime.date)):
                        continue

                    qty = clean_int(qty_val)
                    code = str(code_val).strip() if code_val else ""
                    desc = str(desc_val).strip() if desc_val else ""

                    if qty <= 0 or qty > 1000000 or not code:
                        continue

                    # Skip obvious payment/summary terms or date-like codes
                    lower_code = code.lower()
                    lower_desc = desc.lower()
                    if any(x in lower_code or x in lower_desc for x in ["deposit", "subtotal", "vat", "total due", "payment", "bank details"]):
                        continue
                    if re.match(r"^\d{4}-\d{2}-\d{2}", code) or re.match(r"^\d{2}/\d{2}/\d{4}", code):
                        continue

                    unit_price_val = sheet.cell(row=row, column=col_indices.get("price", 5)).value
                    cost_val = sheet.cell(row=row, column=col_indices.get("cost", 8)).value

                    unit_price = clean_number(unit_price_val)
                    cost_price = clean_number(cost_val)

                    boq_item = BOQItem(
                        boq_id=boq.id,
                        product_code=code,
                        quantity=qty,
                        spec_notes=desc,
                        cost_price=cost_price,
                        retail_price=unit_price
                    )
                    db.add(boq_item)
                    items_count += 1
                except Exception as item_err:
                    # Catch any unexpected cell reading issues
                    pass

            db.commit()
            print(f"  -> Successfully imported {items_count} BOQ line items.")
        except Exception as sheet_err:
            print(f"  -> Error on sheet '{title}': {sheet_err}")
            if db:
                try:
                    db.rollback()
                except Exception:
                    pass
        finally:
            if db:
                db.close()

    os.unlink(tmp_file.name)
    print("\nGoogle Sheets import process completed successfully!")

if __name__ == "__main__":
    run()
