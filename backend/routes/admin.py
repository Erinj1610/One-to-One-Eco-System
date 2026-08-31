from fastapi import APIRouter, Request, HTTPException, UploadFile, File, Depends, Response
from fastapi.responses import FileResponse, JSONResponse
import os
import shutil
from database.cloud_sql import get_db
from sqlalchemy.orm import Session
from models.orm_models import TemplateConfig

router = APIRouter()

# Base path for all system templates (Deployment Trigger Comment)
TEMPLATES_BASE_DIR = os.path.join(os.path.dirname(__file__), '..', 'templates')

def get_template_path(doc_type: str):
    return os.path.abspath(os.path.join(TEMPLATES_BASE_DIR, doc_type, 'template.docx'))

@router.get("/templates/{doc_type}/download")
async def download_template(doc_type: str, db: Session = Depends(get_db)):
    """
    Downloads the current .docx template for the specified document type.
    First checks the database. If not found, falls back to the filesystem.
    If filesystem doesn't exist, it auto-initializes it with a starter template from DESIGN_FEE_PROPOSAL.
    """
    # 1. Check database first
    config = db.query(TemplateConfig).filter(TemplateConfig.template_key == doc_type).first()
    if config and config.docx_binary:
        binary_data = bytes(config.docx_binary)
        if binary_data.startswith(b"UEsDBB") or binary_data.startswith(b"UEsDBQ"):
            import base64
            try:
                binary_data = base64.b64decode(binary_data)
            except Exception:
                pass
        return Response(
            content=binary_data,
            media_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            headers={"Content-Disposition": f"attachment; filename={doc_type.lower()}_template.docx"}
        )

    # 2. Fall back to local file
    path = get_template_path(doc_type)
    if not os.path.exists(path):
        starter_path = get_template_path("DESIGN_FEE_PROPOSAL")
        if os.path.exists(starter_path):
            os.makedirs(os.path.dirname(path), exist_ok=True)
            shutil.copyfile(starter_path, path)
            print(f"DEBUG: Auto-initialized template for {doc_type} from DESIGN_FEE_PROPOSAL")
        else:
            raise HTTPException(status_code=404, detail=f"Template for {doc_type} not found and starter template is missing")
    
    return FileResponse(
        path, 
        media_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        filename=f"{doc_type.lower()}_template.docx"
    )

@router.get("/templates/{doc_type}/xlsx/download")
async def download_xlsx_template(doc_type: str, db: Session = Depends(get_db)):
    """
    Downloads the current .xlsx template for the specified document type.
    First checks the database. If not found, falls back to the filesystem.
    """
    # 1. Check database first
    config = db.query(TemplateConfig).filter(TemplateConfig.template_key == doc_type).first()
    if config and config.xlsx_binary:
        binary_data = bytes(config.xlsx_binary)
        if binary_data.startswith(b"UEsDBB") or binary_data.startswith(b"UEsDBQ"):
            import base64
            try:
                binary_data = base64.b64decode(binary_data)
            except Exception:
                pass
        return Response(
            content=binary_data,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": f"attachment; filename={doc_type.lower()}_template.xlsx"}
        )

    # 2. Fall back to local file
    path = os.path.abspath(os.path.join(TEMPLATES_BASE_DIR, doc_type, 'template.xlsx'))
    if not os.path.exists(path):
        # Auto-create a default Excel layout structure if it doesn't exist on disk
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Template Definition"
        
        ws['A1'] = "CLIENT / CUSTOMER: {{CLIENT_COMPANY}}"
        ws['A2'] = "PROJECT NAME: {{PROJECT_NAME}}"
        ws['A3'] = "DOCUMENT REFERENCE: {{DOCUMENT_NUMBER}}"
        ws['A4'] = "DATE: {{DATE}}"
        
        ws['A6'] = "{{#each items}}"
        ws['A7'] = "{{index}}"
        ws['B7'] = "{{type}}"
        ws['C7'] = "{{description}}"
        ws['D7'] = "{{qty}}"
        ws['E7'] = "{{unitCost}}"
        ws['F7'] = "{{retail}}"
        ws['G7'] = "{{totalRetail}}"
        ws['H7'] = "{{stockStatus}}"
        ws['A8'] = "{{/each}}"
        
        ws['F10'] = "Subtotal:"
        ws['G10'] = "{{SUBTOTAL}}"
        ws['F11'] = "Discount Amount:"
        ws['G11'] = "{{DISCOUNT_AMOUNT}}"
        ws['F12'] = "VAT (15%):"
        ws['G12'] = "{{VAT_AMOUNT}}"
        ws['F13'] = "Grand Total:"
        ws['G13'] = "{{TOTAL_RETAIL}}"
        
        os.makedirs(os.path.dirname(path), exist_ok=True)
        wb.save(path)
        print(f"DEBUG: Auto-initialized Excel template on disk for {doc_type}")
    
    return FileResponse(
        path, 
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        filename=f"{doc_type.lower()}_template.xlsx"
    )

@router.post("/templates/master_excel/upload")
async def upload_master_excel_template(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Uploads the single Master Excel Workbook (.xlsx) containing all document tabs.
    Stores directly in PostgreSQL DB so live deployment always uses the uploaded master template.
    """
    if not file.filename.lower().endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Only .xlsx files are allowed for Master Workbook")
    
    contents = await file.read()
    
    config = db.query(TemplateConfig).filter(TemplateConfig.template_key == "MASTER_EXCEL").first()
    if config:
        config.xlsx_binary = contents
    else:
        config = TemplateConfig(template_key="MASTER_EXCEL", xlsx_binary=contents, config_json={})
        db.add(config)
    
    db.commit()
    return {"message": "Master Excel Workbook uploaded successfully to database"}

@router.post("/templates/master_google_sheet/url")
async def save_master_google_sheet_url(request: Request, db: Session = Depends(get_db)):
    """
    Saves live Master Google Sheet URLs (supports template_key: MASTER_DESIGN_FEE_SHEET, MASTER_ORDERS_SHEET, MASTER_GOOGLE_SHEET).
    """
    data = await request.json()
    url = data.get("url", "").strip()
    key = data.get("key", "MASTER_GOOGLE_SHEET").strip()
    
    config = db.query(TemplateConfig).filter(TemplateConfig.template_key == key).first()
    if config:
        cfg = config.config_json or {}
        cfg["url"] = url
        config.config_json = cfg
    else:
        config = TemplateConfig(template_key=key, config_json={"url": url})
        db.add(config)
    db.commit()
    return {"message": f"Master Google Sheet URL saved for {key}", "key": key, "url": url}

@router.get("/templates/master_google_sheet/url")
async def get_master_google_sheet_url(key: str = "MASTER_GOOGLE_SHEET", db: Session = Depends(get_db)):
    """
    Gets saved Master Google Sheet Spreadsheet URLs by key.
    """
    config = db.query(TemplateConfig).filter(TemplateConfig.template_key == key).first()
    url = (config.config_json or {}).get("url", "") if config else ""
    return {"key": key, "url": url}

@router.post("/templates/{doc_type}/xlsx/upload")
async def upload_xlsx_template(doc_type: str, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Uploads a new .xlsx template for a specific document type and stores it in the database across all alias keys.
    """
    if not file.filename.lower().endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Only .xlsx files are allowed")
    
    contents = await file.read()
    
    alias_keys = [doc_type, doc_type.upper(), doc_type.lower()]
    if doc_type.lower() in ['boq_doc', 'boq']:
        alias_keys.extend(['boq_doc', 'BOQ', 'boq'])
    elif doc_type.lower() in ['schedule', 'lighting_schedule', 'lighting schedule']:
        alias_keys.extend(['schedule', 'SCHEDULE', 'LIGHTING_SCHEDULE'])
    elif doc_type.lower() in ['quote', 'quotation']:
        alias_keys.extend(['quote', 'QUOTATION', 'QUOTE'])
    alias_keys = list(dict.fromkeys(alias_keys))

    for k in alias_keys:
        config = db.query(TemplateConfig).filter(TemplateConfig.template_key == k).first()
        if config:
            config.xlsx_binary = contents
        else:
            config = TemplateConfig(template_key=k, xlsx_binary=contents, config_json={})
            db.add(config)
    
    db.commit()
    return {"message": f"Excel Template for {doc_type} uploaded successfully to database"}

@router.post("/templates/{doc_type}/upload")
async def upload_template(doc_type: str, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Uploads a new .docx template for a specific document type and stores it in the database.
    """
    if not file.filename.lower().endswith('.docx'):
        raise HTTPException(status_code=400, detail="Only .docx files are allowed")
    
    contents = await file.read()
    
    config = db.query(TemplateConfig).filter(TemplateConfig.template_key == doc_type).first()
    if config:
        config.docx_binary = contents
    else:
        config = TemplateConfig(template_key=doc_type, docx_binary=contents, config_json={})
        db.add(config)
    
    db.commit()
    return {"message": f"Template for {doc_type} uploaded successfully to database"}

# --- TEMPLATE CONFIGURATION ENDPOINTS (NO-CODE) ---

@router.get("/configs/{template_key}")
async def get_template_config(template_key: str, db: Session = Depends(get_db)):
    """
    Returns the visual/text configuration for a specific template.
    """
    config = db.query(TemplateConfig).filter(TemplateConfig.template_key == template_key).first()
    if not config:
        # Provide clean, professional defaults if nothing exists
        return {
            "template_key": template_key,
            "config_json": {
                "company_name": "ONE TO ONE LIGHTING",
                "tagline": "Premium Lighting Design & Solutions",
                "address": "123 Solar Street, Cape Town, 8001",
                "contact": "+27 21 000 0000 | hello@onetoone.co.za",
                "terms": "1. This proposal is valid for 30 days from date of issue.\n2. 50% deposit is required before concept design begins.\n3. All designs remain the property of One to One Lighting until full payment.",
                "footer": "One to One Lighting Design | Confidential Proposal",
                "color_theme": "#10b981"
            }
        }
    return {
        "id": config.id,
        "template_key": config.template_key,
        "config_json": config.config_json,
        "html_content": config.html_content
    }

@router.post("/configs/{template_key}")
async def save_template_config(template_key: str, request: Request, db: Session = Depends(get_db)):
    """
    Saves visual/text configuration for a specific template.
    """
    data = await request.json()
    config = db.query(TemplateConfig).filter(TemplateConfig.template_key == template_key).first()
    
    if config:
        config.config_json = data
    else:
        config = TemplateConfig(template_key=template_key, config_json=data)
        db.add(config)
    
    db.commit()
    return {"message": "Design settings saved successfully"}

@router.get("/templates/{doc_type}/html")
async def get_template_html(doc_type: str, db: Session = Depends(get_db)):
    """
    Returns the visually designed HTML template content for a document type.
    """
    config = db.query(TemplateConfig).filter(TemplateConfig.template_key == doc_type).first()
    if not config or not config.html_content:
        return {"html": ""}
    return {"html": config.html_content}

@router.post("/templates/{doc_type}/html")
async def save_template_html(doc_type: str, request: Request, db: Session = Depends(get_db)):
    """
    Saves the visually designed HTML template content for a document type.
    """
    data = await request.json()
    html_val = data.get("html")
    config = db.query(TemplateConfig).filter(TemplateConfig.template_key == doc_type).first()
    
    if config:
        config.html_content = html_val
    else:
        config = TemplateConfig(template_key=doc_type, html_content=html_val, config_json={})
        db.add(config)
        
    db.commit()
    return {"message": "Visual HTML template saved successfully"}


@router.get("/templates/{doc_type}/metadata")
async def get_template_metadata(doc_type: str, db: Session = Depends(get_db)):
    """
    Returns metadata about specified template.
    """
    # 1. Check database first
    config = db.query(TemplateConfig).filter(TemplateConfig.template_key == doc_type).first()
    if config and config.docx_binary:
        binary_data = bytes(config.docx_binary)
        if binary_data.startswith(b"UEsDBB") or binary_data.startswith(b"UEsDBQ"):
            import base64
            try:
                binary_data = base64.b64decode(binary_data)
            except Exception:
                pass
        return {
            "exists": True,
            "size": len(binary_data),
            "last_modified": None
        }
        
    # 2. Fall back to local file
    path = get_template_path(doc_type)
    if not os.path.exists(path):
        return {"exists": False}
    
    stats = os.stat(path)
    return {
        "exists": True,
        "size": stats.st_size,
        "last_modified": stats.st_mtime
    }


import tempfile
from fastapi import Body

@router.post("/generate/{doc_type}")
def generate_document(doc_type: str, page: int = None, format: str = 'pdf', is_save_action: bool = False, data: dict = Body(...), db: Session = Depends(get_db)):
    """
    100% Google Sheets Document Generator Engine.
    Supports Dual Master Templates: MASTER_DESIGN_FEE_SHEET for Design Fees & MASTER_ORDERS_SHEET for Orders/BOQ.
    Pass is_save_action=true to trigger Drive vault revision archival.
    """
    print(f"DEBUG: Generating {doc_type} via Master Google Sheets Engine with tokens: {list(data.keys())} (is_save_action={is_save_action})")
    
    # 1. Determine specific template key based on document type
    target_key = "MASTER_DESIGN_FEE_SHEET" if doc_type == "DESIGN_FEE_PROPOSAL" else "MASTER_ORDERS_SHEET"
    
    specific_config = db.query(TemplateConfig).filter(TemplateConfig.template_key == target_key).first()
    master_gsheet_url = (specific_config.config_json or {}).get("url", "").strip() if specific_config else ""
    
    # Fallback to general MASTER_GOOGLE_SHEET if specific key is empty
    if not master_gsheet_url:
        fallback_config = db.query(TemplateConfig).filter(TemplateConfig.template_key == "MASTER_GOOGLE_SHEET").first()
        master_gsheet_url = (fallback_config.config_json or {}).get("url", "").strip() if fallback_config else ""

    if not master_gsheet_url:
        raise HTTPException(
            status_code=400, 
            detail=f"Master Google Sheet URL for {target_key} is missing. Please set your live Google Sheet links under Settings > Templates."
        )

    # Resolve custom configured Google Sheet tab name for this specific document type
    doc_config = db.query(TemplateConfig).filter(
        or_(
            TemplateConfig.template_key == doc_type,
            TemplateConfig.template_key == doc_type.upper(),
            TemplateConfig.template_key == doc_type.lower()
        )
    ).first()
    configured_tab_name = ""
    if doc_config and doc_config.config_json:
        configured_tab_name = str(doc_config.config_json.get("excel_tab_name") or doc_config.config_json.get("tab_name") or "").strip()
    
    effective_sheet_name = configured_tab_name or doc_type

    service_account_config = db.query(TemplateConfig).filter(TemplateConfig.template_key == "GOOGLE_SERVICE_ACCOUNT_JSON").first()
    credentials_json = (service_account_config.config_json or {}) if service_account_config else None

    try:
        from services.google_doc_engine import merge_google_sheet
        pdf_path, sheet_id, sheet_url = merge_google_sheet(
            template_source=master_gsheet_url,
            tokens=data,
            sheet_name=effective_sheet_name,
            output_pdf_name=f"{doc_type.lower()}.pdf",
            credentials_json=credentials_json,
            is_save_action=is_save_action
        )
        return FileResponse(
            pdf_path,
            media_type='application/pdf',
            filename=f"Proposal_{doc_type.lower()}.pdf",
            headers={"X-Google-Sheet-Url": sheet_url or ""}
        )
    except Exception as gsheet_err:
        print(f"Master Google Sheet Merge Error: {gsheet_err}")
        raise HTTPException(status_code=500, detail=f"Google Sheets PDF Generation Error: {gsheet_err}")


@router.post("/generate-batch")
def generate_batch_documents(request_body: dict = Body(...), db: Session = Depends(get_db)):
    """
    Generates multiple document types sequentially and merges them into a single PDF download stream.
    """
    doc_types = request_body.get('doc_types', [])
    data = request_body.get('data', {})
    is_save_action = request_body.get('is_save_action', False)

    if not doc_types:
        raise HTTPException(status_code=400, detail="No document types specified for batch PDF generation.")

    service_account_config = db.query(TemplateConfig).filter(TemplateConfig.template_key == "GOOGLE_SERVICE_ACCOUNT_JSON").first()
    credentials_json = (service_account_config.config_json or {}) if service_account_config else None

    from services.google_doc_engine import merge_google_sheet
    generated_pdf_paths = []

    for dt in doc_types:
        target_key = "MASTER_DESIGN_FEE_SHEET" if dt == "DESIGN_FEE_PROPOSAL" else "MASTER_ORDERS_SHEET"
        specific_config = db.query(TemplateConfig).filter(TemplateConfig.template_key == target_key).first()
        master_gsheet_url = (specific_config.config_json or {}).get("url", "").strip() if specific_config else ""
        if not master_gsheet_url:
            fallback_config = db.query(TemplateConfig).filter(TemplateConfig.template_key == "MASTER_GOOGLE_SHEET").first()
            master_gsheet_url = (fallback_config.config_json or {}).get("url", "").strip() if fallback_config else ""

        if not master_gsheet_url:
            continue

        doc_config = db.query(TemplateConfig).filter(
            or_(
                TemplateConfig.template_key == dt,
                TemplateConfig.template_key == dt.upper(),
                TemplateConfig.template_key == dt.lower()
            )
        ).first()
        configured_tab_name = ""
        if doc_config and doc_config.config_json:
            configured_tab_name = str(doc_config.config_json.get("excel_tab_name") or doc_config.config_json.get("tab_name") or "").strip()
        effective_sheet_name = configured_tab_name or dt

        try:
            pdf_p, _, _ = merge_google_sheet(
                template_source=master_gsheet_url,
                tokens=data,
                sheet_name=effective_sheet_name,
                output_pdf_name=f"{dt.lower()}.pdf",
                credentials_json=credentials_json,
                is_save_action=is_save_action
            )
            if pdf_p and os.path.exists(pdf_p):
                generated_pdf_paths.append(pdf_p)
        except Exception as err:
            print(f"Error generating {dt} for batch PDF: {err}")

    if not generated_pdf_paths:
        raise HTTPException(status_code=500, detail="Failed to generate any of the selected document PDFs.")

    if len(generated_pdf_paths) == 1:
        return FileResponse(
            generated_pdf_paths[0],
            media_type='application/pdf',
            filename="Combined_Documents.pdf"
        )

    # Merge multiple PDFs into a single file
    try:
        import pypdf
        merger = pypdf.PdfWriter()
    except ImportError:
        try:
            import PyPDF2 as pypdf
            merger = pypdf.PdfMerger()
        except ImportError:
            raise HTTPException(status_code=500, detail="PDF merger library (pypdf) is missing on server.")

    merged_temp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    merged_path = merged_temp.name
    merged_temp.close()

    try:
        for p in generated_pdf_paths:
            merger.append(p)
        merger.write(merged_path)
        merger.close()

        return FileResponse(
            merged_path,
            media_type='application/pdf',
            filename="Combined_Documents.pdf"
        )
    except Exception as merge_err:
        print(f"Error merging PDFs: {merge_err}")
        # Fallback to returning first PDF
        return FileResponse(
            generated_pdf_paths[0],
            media_type='application/pdf',
            filename="Combined_Documents.pdf"
        )

    # 2. Check if a direct docx template exists (either in DB or on disk)
    from services.docx_engine import merge_docx_template
    
    custom_template_temp_path = None
    temp_dir = None
    docx_template_path = None
    
    if config and config.docx_binary:
        print(f"DEBUG: Using custom template from database for {doc_type}")
        temp_dir = tempfile.mkdtemp()
        custom_template_temp_path = os.path.join(temp_dir, "db_template.docx")
        
        # PostgreSQL might return the docx file encoded as a base64 string bytes or base64 string
        binary_data = bytes(config.docx_binary)
        if binary_data.startswith(b"UEsDBB") or binary_data.startswith(b"UEsDBQ"):
            import base64
            try:
                binary_data = base64.b64decode(binary_data)
                print("DEBUG: Successfully base64 decoded custom template from DB")
            except Exception as b64_err:
                print(f"DEBUG: Failed to base64 decode custom template: {b64_err}")
                
        with open(custom_template_temp_path, "wb") as f:
            f.write(binary_data)
        docx_template_path = custom_template_temp_path
    else:
        disk_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'templates', doc_type, 'template.docx'))
        if os.path.exists(disk_path):
            docx_template_path = disk_path
    
    if docx_template_path:
        print(f"DEBUG: Found direct docx template at {docx_template_path}. Using docx_engine...")
        custom_creds = None
        if config:
            custom_creds = config.config_json.get("google_credentials_json")
            
        if not custom_creds:
            all_configs = db.query(TemplateConfig).all()
            for ac in all_configs:
                if ac.config_json and ac.config_json.get("google_credentials_json"):
                    custom_creds = ac.config_json.get("google_credentials_json")
                    print(f"DEBUG: Falling back to credentials from config '{ac.template_key}' for docx merge")
                    break

        if custom_creds and isinstance(custom_creds, str):
            import json
            try:
                custom_creds = json.loads(custom_creds)
            except Exception as j_err:
                print(f"DEBUG: JSON credentials load error: {j_err}")
                custom_creds = None
                    
        doc_id_to_use = config.config_json.get("google_doc_id") if config else None
        if not doc_id_to_use:
            all_configs = db.query(TemplateConfig).all()
            for ac in all_configs:
                if ac.config_json and ac.config_json.get("google_doc_id"):
                    doc_id_to_use = ac.config_json.get("google_doc_id")
                    print(f"DEBUG: Falling back to google_doc_id '{doc_id_to_use}' from config {ac.template_key}")
                    break

        try:
            pdf_path = merge_docx_template(
                docx_template_path,
                data,
                f"{doc_type.lower()}.pdf",
                credentials_json=custom_creds,
                google_doc_id=doc_id_to_use
            )
            print(f"DEBUG: Generation successful from docx! PDF path: {pdf_path}")
            
            if page is not None:
                import pypdf
                try:
                    reader = pypdf.PdfReader(pdf_path)
                    total_pages = len(reader.pages)
                    idx = max(0, min(page - 1, total_pages - 1))
                    
                    writer = pypdf.PdfWriter()
                    writer.add_page(reader.pages[idx])
                    
                    single_page_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
                    single_page_pdf_path = single_page_pdf.name
                    single_page_pdf.close()
                    
                    with open(single_page_pdf_path, "wb") as f:
                        writer.write(f)
                    
                    # Remove original full PDF
                    try:
                        os.remove(pdf_path)
                    except Exception:
                        pass
                    pdf_path = single_page_pdf_path
                    print(f"DEBUG: Successfully extracted page {page} to {pdf_path}")
                except Exception as pypdf_err:
                    print(f"Error extracting page {page} with pypdf: {pypdf_err}")

            
            filename = f"Document_{doc_type.lower()}.pdf"
            if config:
                naming_conv = config.config_json.get("naming_convention")
                if naming_conv:
                    temp_name = naming_conv
                    for k, v in data.items():
                        temp_name = temp_name.replace("{{" + k + "}}", str(v))
                    import re
                    filename = re.sub(r'[\\/*?:"<>|]', "", temp_name)
                    if not filename.lower().endswith(".pdf"):
                        filename += ".pdf"
                        
            return FileResponse(
                pdf_path,
                media_type='application/pdf',
                filename=filename
            )
        except Exception as docx_err:
            print(f"Error generating {doc_type} via docx: {docx_err}")
            raise HTTPException(status_code=500, detail=f"Word Template Conversion Error: {docx_err}")
        finally:
            # Clean up the custom template temp files if they were written
            if custom_template_temp_path and os.path.exists(custom_template_temp_path):
                try:
                    os.remove(custom_template_temp_path)
                except Exception:
                    pass
            if temp_dir and os.path.exists(temp_dir):
                try:
                    os.rmdir(temp_dir)
                except Exception:
                    pass

    # Fetch template settings to get the linked Google Doc ID and Credentials
    config = db.query(TemplateConfig).filter(TemplateConfig.template_key == doc_type).first()
    google_doc_id = None
    if config and config.config_json and "google_doc_id" in config.config_json:
        google_doc_id = config.config_json["google_doc_id"]
    else:
        print(f"DEBUG: No google_doc_id in template config found for {doc_type}, using default fallback")
        google_doc_id = "1E6tnSk6jxXUM100lVJDUb_7ezYDQwo8TBpL975tr6Og"

    print(f"DEBUG: Using Google Doc ID: {google_doc_id}")
    
    # Extract credentials if provided manually in the Hub
    custom_creds = config.config_json.get("google_credentials_json") if (config and config.config_json) else None
    if not custom_creds:
        all_configs = db.query(TemplateConfig).all()
        for ac in all_configs:
            if ac.config_json and ac.config_json.get("google_credentials_json"):
                custom_creds = ac.config_json.get("google_credentials_json")
                print(f"DEBUG: Falling back to credentials from config '{ac.template_key}' for Google Doc merge")
                break

    if custom_creds:
        print("DEBUG: Private Service Account JSON detected. Attempting to parse...")
        if isinstance(custom_creds, str):
            import json
            try:
                custom_creds = json.loads(custom_creds)
                print("DEBUG: JSON credentials parsed successfully.")
            except Exception as j_err:
                print(f"DEBUG: JSON parse error: {j_err}")
                custom_creds = None
    else:
        print("DEBUG: No manual JSON credentials found. Falling back to Application Default.")

    try:
        # Perform the merge using our professional service (Sync)
        pdf_path = merge_google_doc(
            google_doc_id, 
            data, 
            f"{doc_type.lower()}.pdf", 
            credentials_json=custom_creds
        )
        print(f"DEBUG: Generation successful! PDF path: {pdf_path}")

        # Handle naming convention
        filename = f"Proposal_{doc_type.lower()}.pdf"
        naming_conv = config.config_json.get("naming_convention") if (config and config.config_json) else None
        if naming_conv:
            # Simple replacement for the filename
            temp_name = naming_conv
            for k, v in data.items():
                temp_name = temp_name.replace("{{" + k + "}}", str(v))
            
            # Clean filename (remove problematic chars)
            import re
            filename = re.sub(r'[\\/*?:"<>|]', "", temp_name)
            if not filename.lower().endswith(".pdf"):
                filename += ".pdf"

        return FileResponse(
            pdf_path, 
            media_type='application/pdf', 
            filename=filename
        )
            
    except Exception as e:
        error_str = str(e)
        print(f"Error generating {doc_type}: {error_str}")
        
        # User-Friendly Error Mapping
        if "403" in error_str or "forbidden" in error_str.lower():
            friendly_detail = "Permission Denied: Please ensure the Google Doc is SHARED with your service account email as a 'Viewer'."
        elif "404" in error_str or "not found" in error_str.lower():
            friendly_detail = "Document Not Found: Please verify the Google Doc ID/URL in the Branding Hub."
        elif "API has not been used" in error_str:
            friendly_detail = "API Not Enabled: Please ensure Google Docs and Google Drive APIs are ENABLED in your GCP Console."
        else:
            friendly_detail = f"Google Docs Error: {error_str}"

        raise HTTPException(status_code=500, detail=friendly_detail)

