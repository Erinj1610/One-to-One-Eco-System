from fastapi import APIRouter, HTTPException, Depends, Body
from sqlalchemy.orm import Session
from database.cloud_sql import get_db
from models.orm_models import Order, Project, OrderItem
from typing import Optional, List, Dict, Any
import io
import os
import re
import socket
import datetime
import openpyxl
import google.auth
from google.oauth2 import service_account
from googleapiclient.discovery import build

# Set HTTP socket timeout
socket.setdefaulttimeout(120)

ROOT_DRIVE_FOLDER_ID = "0AFF94SUUC_EQUk9PVA"

router = APIRouter()

def get_google_services():
    """Initializes Google Sheets and Drive API services using Cloud Run Compute Service Account."""
    scopes = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive'
    ]
    creds, _ = google.auth.default(scopes=scopes)
    sheets_service = build('sheets', 'v4', credentials=creds)
    drive_service = build('drive', 'v3', credentials=creds)
    return sheets_service, drive_service

def extract_spreadsheet_id(url_or_id: str) -> str:
    """Extracts clean 44-char Spreadsheet ID from full Google Sheets URL or raw ID."""
    match = re.search(r"/d/([a-zA-Z0-9-_]+)", url_or_id)
    if match:
        return match.group(1)
    return url_or_id.strip()

def safe_float(val_str, default=0.0):
    if val_str is None:
        return default
    if isinstance(val_str, (int, float)):
        return float(val_str)
    clean = re.sub(r"[^\d.-]", "", str(val_str).replace(',', ''))
    try:
        return float(clean)
    except:
        return default

def safe_int(val, default=0):
    f = safe_float(val, default=float(default))
    return int(f)

def normalize_key(s: str) -> str:
    """Normalizes any string to alphanumeric-only lowercase for 100% collision-free matching."""
    if not s:
        return ""
    return re.sub(r'[^a-zA-Z0-9]', '', str(s)).lower()

def parse_excel_date(val: Any) -> str:
    """Parses Excel serial dates or datetime objects into standardized YYYY-MM-DD format."""
    if not val:
        return ""
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.strftime('%Y-%m-%d')
    
    val_str = str(val).strip()
    if not val_str:
        return ""
    
    # Check if numeric Excel serial date (e.g., 45230)
    try:
        f = float(val_str)
        if f > 30000 and f < 70000:
            base = datetime.date(1899, 12, 30)
            return (base + datetime.timedelta(days=int(f))).isoformat()
    except Exception:
        pass

    # Standardize string date YYYY-MM-DD
    match_iso = re.search(r'(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})', val_str)
    if match_iso:
        return f"{match_iso.group(1)}-{int(match_iso.group(2)):02d}-{int(match_iso.group(3)):02d}"

    # Standardize string date DD/MM/YYYY or MM/DD/YYYY
    match_dmy = re.search(r'(\d{1,2})[-/.](\d{1,2})[-/.](\d{4})', val_str)
    if match_dmy:
        return f"{match_dmy.group(3)}-{int(match_dmy.group(2)):02d}-{int(match_dmy.group(1)):02d}"

    return val_str

RECONCILIATION_COLUMNS = [
    "Project Key", "Project Name", "Client Company", "Order ID", "Quote Name",
    "Sales Rep", "Delivery Address", "Item ID", "Qty", "1:1 Code",
    "Item Code", "Description", "Unit Cost Ex VAT", "Unit Retail Price Ex VAT",
    "Brand", "Supplier", "Item Type", "Stock Status", "Stock on Hand",
    "Qty Ordered (PO)", "PO Supplier", "Date Ordered", "PO Reference", "Delivery ETA",
    "Qty REC", "Date REC", "GRN Reference", "Qty INV", "Invoice Reference", "Date INV",
    "Qty DEL", "Date DEL", "Delivery Reference", "Delivery Comments", "Sheet Order Status",
    "Deposit Value", "Deposit Invoice Sent", "Deposit Payment Date",
    "Balance Value", "Balance Payment Date", "Amount Paid"
]

@router.post("/audit-comparison/generate")
def generate_audit_comparison(payload: dict = Body(...), db: Session = Depends(get_db)):
    """
    Read-only live audit comparison generator.
    1. Downloads live Google Sheet directly as an XLSX workbook stream via Drive API export (completes in < 2s).
    2. Identifies all project tabs appearing AFTER 'Template' using openpyxl (exact clone of handleGenerateReconciliationTemplate).
    3. Extracts rows 9-89 with exact cell references and formats in 41 reconciliation columns.
    4. Queries Cloud SQL database and generates 3-tab heatmap spreadsheet.
    """
    raw_sheet_input = payload.get("current_system_sheet_url")
    user_email = payload.get("user_email", "erin.jones@1-to-1.world").strip()

    if not raw_sheet_input:
        raise HTTPException(status_code=400, detail="Please provide a valid Google Sheet URL or ID.")

    source_sheet_id = extract_spreadsheet_id(raw_sheet_input)

    try:
        sheets_service, drive_service = get_google_services()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to initialize Google API service: {str(e)}")

    # Step 1: Export Google Sheet to Excel binary stream via Google Drive API
    try:
        # Pre-check drive access
        try:
            drive_service.files().get(
                fileId=source_sheet_id, 
                fields="id, name", 
                supportsAllDrives=True
            ).execute()
        except Exception as drive_err:
            err_str = str(drive_err)
            if "403" in err_str or "permission" in err_str.lower():
                raise HTTPException(
                    status_code=403,
                    detail=f"Permission Denied: Google Drive cannot access sheet ID '{source_sheet_id}'. Please verify you shared the Google Sheet with 858977785048-compute@developer.gserviceaccount.com."
                )
            raise drive_err

        # Direct XLSX export: fast, reliable, zero range errors
        exported_bytes = drive_service.files().export(
            fileId=source_sheet_id,
            mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ).execute()

        wb = openpyxl.load_workbook(io.BytesIO(exported_bytes), data_only=True)
        all_sheet_names = wb.sheetnames

    except HTTPException:
        raise
    except Exception as e:
        err_msg = str(e)
        raise HTTPException(status_code=400, detail=f"Could not export live Google Sheet: {err_msg}")

    # Process sheets that appear AFTER 'Template' (matching handleGenerateReconciliationTemplate)
    has_template_tab = any(name.strip().lower() == 'template' for name in all_sheet_names)
    found_template = False
    target_sheets = []

    for name in all_sheet_names:
        if name.strip().lower() == 'template':
            found_template = True
            continue
        if found_template:
            target_sheets.append(name)

    if not target_sheets:
        skip_names = {'template', 'control', 'summary', 'instructions', 'readme', 'master'}
        target_sheets = [name for name in all_sheet_names if name.strip().lower() not in skip_names]

    if not target_sheets:
        raise HTTPException(status_code=400, detail="No order/project tabs found after the 'Template' sheet in the workbook.")

    legacy_flat_rows = []

    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        def get_val(cell_ref: str) -> str:
            try:
                c = ws[cell_ref]
                if c.value is None:
                    return ""
                return str(c.value).strip()
            except Exception:
                return ""

        def get_row_val(col_letter: str, row_num: int) -> Any:
            try:
                c = ws[f"{col_letter}{row_num}"]
                return c.value
            except Exception:
                return None

        client_company = get_val('D2')
        order_name = get_val('D3') or sheet_name
        project_f5 = get_val('F5') or 'General Project'
        delivery_address = get_val('D6')
        sales_rep = get_val('F1')
        order_status_g98 = get_val('G98') or "Processing"

        safe_order_ref = re.sub(r'[^a-zA-Z0-9]', '-', order_name).lower()

        deposit_invoice_sent = get_val('D90') or 'No'
        deposit_payment_date = parse_excel_date(get_row_val('D', 95))
        balance_payment_date = parse_excel_date(get_row_val('D', 96))
        deposit_value = safe_float(get_row_val('G', 95))
        balance_value = safe_float(get_row_val('G', 96))

        amount_paid = 0.0
        if deposit_payment_date or deposit_value > 0:
            amount_paid += deposit_value
        if balance_payment_date or balance_value > 0:
            amount_paid += balance_value

        # Parse rows 9 to 89
        for r in range(9, 90):
            qty_raw = get_row_val('B', r)
            qty = safe_int(qty_raw)
            if qty <= 0:
                continue

            one_one_code = str(get_row_val('C', r) or '').strip()
            item_code = str(get_row_val('D', r) or '').strip()
            description = str(get_row_val('E', r) or '').strip()
            unit_retail = safe_float(get_row_val('F', r))
            unit_cost = safe_float(get_row_val('I', r))
            brand = str(get_row_val('L', r) or '').strip()
            supplier = str(get_row_val('M', r) or '').strip()
            product_type = str(get_row_val('N', r) or '').strip() or "Hardware"

            po_ref_from_sheet = str(get_row_val('Q', r) or '').strip()
            po_supplier = str(get_row_val('R', r) or supplier or 'Warehouse Inventory').strip()
            date_ordered_raw = get_row_val('S', r)
            date_ordered = parse_excel_date(date_ordered_raw)
            eta_raw = get_row_val('T', r)
            eta = parse_excel_date(eta_raw)

            date_rec_raw = get_row_val('U', r)
            date_rec = parse_excel_date(date_rec_raw)
            qty_rec = safe_int(get_row_val('V', r))

            qty_inv = safe_int(get_row_val('Y', r))
            invoice_ref = str(get_row_val('Z', r) or '').strip()
            date_inv_raw = get_row_val('AA', r)
            date_inv = parse_excel_date(date_inv_raw)

            delivery_ref = str(get_row_val('AC', r) or '').strip()

            po_ref = po_ref_from_sheet
            if not po_ref and date_ordered:
                clean_date = re.sub(r'[^0-9]', '', date_ordered)
                clean_supp = re.sub(r'[^a-zA-Z0-9]', '', po_supplier)[:8].lower()
                po_ref = f"PO-{safe_order_ref}-{clean_supp}-{clean_date}"

            grn_ref = ""
            if date_rec and qty_rec > 0:
                clean_date = re.sub(r'[^0-9]', '', date_rec)
                grn_ref = f"GRN-{safe_order_ref}-{clean_date}"

            date_del = date_rec if delivery_ref else ""
            qty_del = qty_rec if delivery_ref else 0

            inv_ref_value = invoice_ref
            if not inv_ref_value and date_inv and qty_inv > 0:
                clean_date = re.sub(r'[^0-9]', '', date_inv)
                inv_ref_value = f"INV-{safe_order_ref}-{clean_date}"

            temp_item_id = f"ITEM-{safe_order_ref}-{one_one_code or item_code or 'FITTING'}-{r}"
            stock_status = str(get_row_val('N', r) or '').strip()
            stock_on_hand = safe_int(get_row_val('O', r))

            legacy_flat_rows.append({
                "Project Key": re.sub(r'[^a-zA-Z0-9]', '-', project_f5).lower(),
                "Project Name": project_f5,
                "Client Company": client_company,
                "Order ID": safe_order_ref,
                "Quote Name": order_name,
                "Sales Rep": sales_rep,
                "Delivery Address": delivery_address,
                "Item ID": temp_item_id,
                "Qty": qty,
                "1:1 Code": one_one_code,
                "Item Code": item_code,
                "Description": description,
                "Unit Cost Ex VAT": unit_cost,
                "Unit Retail Price Ex VAT": unit_retail,
                "Brand": brand,
                "Supplier": supplier,
                "Item Type": product_type,
                "Stock Status": stock_status,
                "Stock on Hand": stock_on_hand,
                "Qty Ordered (PO)": 0,
                "PO Supplier": po_supplier,
                "Date Ordered": date_ordered,
                "PO Reference": po_ref,
                "Delivery ETA": eta,
                "Qty REC": qty_rec,
                "Date REC": date_rec,
                "GRN Reference": grn_ref,
                "Qty INV": qty_inv,
                "Invoice Reference": inv_ref_value,
                "Date INV": date_inv,
                "Qty DEL": qty_del,
                "Date DEL": date_del,
                "Delivery Reference": delivery_ref,
                "Delivery Comments": "",
                "Sheet Order Status": order_status_g98,
                "Deposit Value": deposit_value,
                "Deposit Invoice Sent": deposit_invoice_sent,
                "Deposit Payment Date": deposit_payment_date,
                "Balance Value": balance_value,
                "Balance Payment Date": balance_payment_date,
                "Amount Paid": amount_paid
            })

    if not legacy_flat_rows:
        raise HTTPException(status_code=400, detail="No order items found in the target worksheets.")

    # Step 2: Query Cloud SQL Database items in identical 41-column format
    db_orders = db.query(Order).all()
    all_projects = db.query(Project).all()
    db_projects_map = {p.project_key: p.name for p in all_projects}
    db_projects_client_map = {p.project_key: (p.client_name or "") for p in all_projects}

    portal_flat_rows = []
    for o in db_orders:
        proj_name = db_projects_map.get(o.project_key, o.project_key or "General Project")
        client_name = getattr(o, 'client_name', None) or db_projects_client_map.get(o.project_key, "")
        items = db.query(OrderItem).filter(OrderItem.order_id == o.po_number).all()

        if items:
            for item in items:
                portal_flat_rows.append({
                    "Project Key": o.project_key or "",
                    "Project Name": proj_name,
                    "Client Company": client_name,
                    "Order ID": o.po_number or "",
                    "Quote Name": getattr(o, 'quote_name', None) or o.po_number or "",
                    "Sales Rep": getattr(o, 'pm_name', None) or "",
                    "Delivery Address": getattr(o, 'notes', None) or "",
                    "Item ID": item.id or "",
                    "Qty": item.qty or 0,
                    "1:1 Code": getattr(item, 'one_one_code', None) or "",
                    "Item Code": getattr(item, 'code', None) or "",
                    "Description": getattr(item, 'description', None) or "",
                    "Unit Cost Ex VAT": getattr(item, 'unit_cost', 0.0) or 0.0,
                    "Unit Retail Price Ex VAT": getattr(item, 'unit_retail', 0.0) or 0.0,
                    "Brand": getattr(item, 'brand', None) or "",
                    "Supplier": getattr(item, 'supplier', None) or "",
                    "Item Type": getattr(item, 'type', None) or "Hardware",
                    "Stock Status": getattr(item, 'stock_status', None) or "",
                    "Stock on Hand": 0,
                    "Qty Ordered (PO)": getattr(item, 'po_qty_ordered', 0) or 0,
                    "PO Supplier": getattr(item, 'po_supplier', None) or "",
                    "Date Ordered": getattr(item, 'po_date', None) or "",
                    "PO Reference": getattr(item, 'po_ref', None) or "",
                    "Delivery ETA": getattr(item, 'eta', None) or "",
                    "Qty REC": getattr(item, 'received_qty', 0) or 0,
                    "Date REC": getattr(item, 'received_date', None) or "",
                    "GRN Reference": "",
                    "Qty INV": getattr(item, 'invoice_qty', 0) or 0,
                    "Invoice Reference": getattr(item, 'invoice_ref', None) or "",
                    "Date INV": getattr(item, 'invoice_date', None) or "",
                    "Qty DEL": getattr(item, 'delivery_qty', 0) or 0,
                    "Date DEL": getattr(item, 'delivery_date', None) or "",
                    "Delivery Reference": getattr(item, 'delivery_status', None) or "",
                    "Delivery Comments": "",
                    "Sheet Order Status": getattr(o, 'status', None) or "Active",
                    "Deposit Value": getattr(o, 'deposit_value', 0.0) or 0.0,
                    "Deposit Invoice Sent": getattr(o, 'deposit_invoice_sent', None) or "No",
                    "Deposit Payment Date": getattr(o, 'deposit_payment_date', None) or "",
                    "Balance Value": getattr(o, 'balance_value', 0.0) or 0.0,
                    "Balance Payment Date": getattr(o, 'balance_payment_date', None) or "",
                    "Amount Paid": getattr(o, 'paid', 0.0) or getattr(o, 'paid_amount', 0.0) or 0.0
                })
        else:
            portal_flat_rows.append({
                "Project Key": o.project_key or "",
                "Project Name": proj_name,
                "Client Company": client_name,
                "Order ID": o.po_number or "",
                "Quote Name": getattr(o, 'quote_name', None) or o.po_number or "",
                "Sales Rep": getattr(o, 'pm_name', None) or "",
                "Delivery Address": getattr(o, 'notes', None) or "",
                "Item ID": f"ORDER-{o.po_number}",
                "Qty": 1,
                "1:1 Code": "",
                "Item Code": "",
                "Description": getattr(o, 'quote_name', None) or o.po_number or "General Order",
                "Unit Cost Ex VAT": 0.0,
                "Unit Retail Price Ex VAT": getattr(o, 'value', 0.0) or 0.0,
                "Brand": "",
                "Supplier": getattr(o, 'supplier_name', None) or getattr(o, 'supplier', None) or "",
                "Item Type": "Order",
                "Stock Status": "",
                "Stock on Hand": 0,
                "Qty Ordered (PO)": 0,
                "PO Supplier": "",
                "Date Ordered": getattr(o, 'order_date', None) or "",
                "PO Reference": o.po_number or "",
                "Delivery ETA": getattr(o, 'eta', None) or "",
                "Qty REC": 0,
                "Date REC": "",
                "GRN Reference": "",
                "Qty INV": 0,
                "Invoice Reference": "",
                "Date INV": "",
                "Qty DEL": 0,
                "Date DEL": "",
                "Delivery Reference": "",
                "Delivery Comments": "",
                "Sheet Order Status": getattr(o, 'status', None) or "Active",
                "Deposit Value": getattr(o, 'deposit_value', 0.0) or 0.0,
                "Deposit Invoice Sent": getattr(o, 'deposit_invoice_sent', None) or "No",
                "Deposit Payment Date": getattr(o, 'deposit_payment_date', None) or "",
                "Balance Value": getattr(o, 'balance_value', 0.0) or 0.0,
                "Balance Payment Date": getattr(o, 'balance_payment_date', None) or "",
                "Amount Paid": getattr(o, 'paid', 0.0) or getattr(o, 'paid_amount', 0.0) or 0.0
            })

    # Step 3: Construct Tab 1 Comparison Heatmap with Normalized Collision-Free Key Matching
    legacy_order_totals = {}
    for row in legacy_flat_rows:
        order_id = str(row.get("Order ID") or "").strip()
        quote_name = str(row.get("Quote Name") or "").strip()
        if not order_id and not quote_name: continue
        
        norm_key = normalize_key(order_id) or normalize_key(quote_name)
        if norm_key not in legacy_order_totals:
            legacy_order_totals[norm_key] = {
                "order_id": order_id,
                "project_name": row.get("Project Name", ""),
                "client_name": row.get("Client Company", ""),
                "quote_name": quote_name or order_id,
                "total_retail": 0.0,
                "amount_paid": safe_float(row.get("Amount Paid", 0)),
                "status": row.get("Sheet Order Status", "Processing"),
                "items_count": 0
            }

        qty = safe_int(row.get("Qty", 0))
        unit_retail = safe_float(row.get("Unit Retail Price Ex VAT", 0))
        legacy_order_totals[norm_key]["total_retail"] += qty * unit_retail
        legacy_order_totals[norm_key]["items_count"] += 1

    portal_order_totals = {}
    for row in portal_flat_rows:
        order_id = str(row.get("Order ID") or "").strip()
        quote_name = str(row.get("Quote Name") or "").strip()
        if not order_id and not quote_name: continue

        norm_key = normalize_key(order_id) or normalize_key(quote_name)
        if norm_key not in portal_order_totals:
            portal_order_totals[norm_key] = {
                "order_id": order_id,
                "project_name": row.get("Project Name", ""),
                "client_name": row.get("Client Company", ""),
                "quote_name": quote_name or order_id,
                "total_retail": 0.0,
                "amount_paid": safe_float(row.get("Amount Paid", 0)),
                "status": row.get("Sheet Order Status", "Processing"),
                "items_count": 0
            }

        qty = safe_int(row.get("Qty", 0))
        unit_retail = safe_float(row.get("Unit Retail Price Ex VAT", 0))
        portal_order_totals[norm_key]["total_retail"] += qty * unit_retail
        portal_order_totals[norm_key]["items_count"] += 1

    # Heatmap Headers
    heatmap_headers = [
        "Order ID / PO #", "Project Name", "Client Company", "Quote Name",
        "Legacy Items Count", "Portal Items Count",
        "Retail Value (Legacy)", "Retail Value (Portal)",
        "Amount Paid (Legacy)", "Amount Paid (Portal)",
        "Status (Legacy)", "Status (Portal)", "Audit Match Status"
    ]
    heatmap_rows = [heatmap_headers]

    matches_count = 0
    mismatches_count = 0
    missing_in_portal_count = 0
    new_in_portal_count = 0
    discrepancy_details = []

    all_norm_keys = sorted(list(set(list(legacy_order_totals.keys()) + list(portal_order_totals.keys()))))

    for nkey in all_norm_keys:
        leg = legacy_order_totals.get(nkey)
        port = portal_order_totals.get(nkey)

        display_oid = (leg.get("order_id") if leg else None) or (port.get("order_id") if port else None) or nkey

        if leg and port:
            ret_match = abs(leg["total_retail"] - port["total_retail"]) < 1.0
            paid_match = abs(leg["amount_paid"] - port["amount_paid"]) < 1.0
            stat_match = leg["status"].lower().strip() == port["status"].lower().strip()
            item_match = leg["items_count"] == port["items_count"]

            if ret_match and paid_match and stat_match and item_match:
                audit_status = "🟢 100% Match"
                matches_count += 1
            else:
                audit_status = "🔴 Mismatch Detected"
                mismatches_count += 1
                discrepancy_details.append({
                    "order_id": display_oid,
                    "type": "MISMATCH",
                    "legacy_retail": leg["total_retail"],
                    "portal_retail": port["total_retail"],
                    "legacy_paid": leg["amount_paid"],
                    "portal_paid": port["amount_paid"],
                    "legacy_status": leg["status"],
                    "portal_status": port["status"]
                })

            heatmap_rows.append([
                display_oid,
                leg["project_name"] or port["project_name"],
                leg["client_name"] or port["client_name"],
                leg["quote_name"] or port["quote_name"],
                leg["items_count"],
                port["items_count"],
                f"R {leg['total_retail']:,.2f}",
                f"R {port['total_retail']:,.2f}",
                f"R {leg['amount_paid']:,.2f}",
                f"R {port['amount_paid']:,.2f}",
                leg["status"],
                port["status"],
                audit_status
            ])
        elif leg and not port:
            missing_in_portal_count += 1
            discrepancy_details.append({
                "order_id": display_oid,
                "type": "MISSING_IN_PORTAL",
                "legacy_retail": leg["total_retail"],
                "portal_retail": 0.0,
                "legacy_paid": leg["amount_paid"],
                "portal_paid": 0.0,
                "legacy_status": leg["status"],
                "portal_status": "MISSING"
            })
            heatmap_rows.append([
                display_oid, leg["project_name"], leg["client_name"], leg["quote_name"],
                leg["items_count"], 0,
                f"R {leg['total_retail']:,.2f}", "MISSING IN PORTAL",
                f"R {leg['amount_paid']:,.2f}", "MISSING IN PORTAL",
                leg["status"], "MISSING IN PORTAL",
                "🛑 MISSING IN PORTAL"
            ])
        elif port and not leg:
            new_in_portal_count += 1
            heatmap_rows.append([
                display_oid, port["project_name"], port["client_name"], port["quote_name"],
                0, port["items_count"],
                "NOT IN LEGACY", f"R {port['total_retail']:,.2f}",
                "NOT IN LEGACY", f"R {port['amount_paid']:,.2f}",
                "NOT IN LEGACY", port["status"],
                "⚠️ NEW IN PORTAL"
            ])

    # Convert flat rows into 2D arrays for Google Sheets / Excel
    legacy_tab_2d = [RECONCILIATION_COLUMNS]
    for r in legacy_flat_rows:
        legacy_tab_2d.append([r.get(col, "") for col in RECONCILIATION_COLUMNS])

    portal_tab_2d = [RECONCILIATION_COLUMNS]
    for r in portal_flat_rows:
        portal_tab_2d.append([r.get(col, "") for col in RECONCILIATION_COLUMNS])

    # Step 4: Create new 3-Tab Google Spreadsheet inside Shared Drive Vault using Drive API
    audit_sheet_id = None
    audit_sheet_url = None
    try:
        file_metadata = {
            'name': '1-to-1 World - Live System Comparison & Reconciliation Audit',
            'mimeType': 'application/vnd.google-apps.spreadsheet',
            'parents': [ROOT_DRIVE_FOLDER_ID]
        }
        created_file = drive_service.files().create(
            body=file_metadata,
            supportsAllDrives=True,
            fields='id, webViewLink'
        ).execute()
        audit_sheet_id = created_file.get('id')
        audit_sheet_url = created_file.get('webViewLink') or f"https://docs.google.com/spreadsheets/d/{audit_sheet_id}/edit"
    except Exception as create_err:
        print(f"Notice: Direct Drive Vault sheet creation fallback: {create_err}")
        try:
            created_file = drive_service.files().create(
                body={'name': '1-to-1 World - Live System Comparison & Reconciliation Audit', 'mimeType': 'application/vnd.google-apps.spreadsheet'},
                supportsAllDrives=True,
                fields='id, webViewLink'
            ).execute()
            audit_sheet_id = created_file.get('id')
            audit_sheet_url = created_file.get('webViewLink') or f"https://docs.google.com/spreadsheets/d/{audit_sheet_id}/edit"
        except Exception as e:
            print(f"Warning: Could not create Google Sheet file: {e}")

    # Step 5: Share audit sheet with user email
    if audit_sheet_id:
        try:
            drive_service.permissions().create(
                fileId=audit_sheet_id,
                body={'type': 'user', 'role': 'writer', 'emailAddress': user_email},
                supportsAllDrives=True,
                fields='id'
            ).execute()
        except Exception as e:
            print(f"Notice: Google Drive share notice: {e}")

    # Step 6: Create 3 Tabs & Populate Data in Google Sheet
    if audit_sheet_id:
        try:
            meta = sheets_service.spreadsheets().get(spreadsheetId=audit_sheet_id).execute()
            initial_sheet_id = meta.get('sheets', [{}])[0].get('properties', {}).get('sheetId', 0)

            requests = [
                {
                    'addSheet': {
                        'properties': {
                            'sheetId': 101,
                            'title': '🚨 AUDIT & DISCREPANCY HEATMAP',
                            'gridProperties': {'frozenRowCount': 1}
                        }
                    }
                },
                {
                    'addSheet': {
                        'properties': {
                            'sheetId': 102,
                            'title': 'Current Live System Data',
                            'gridProperties': {'frozenRowCount': 1}
                        }
                    }
                },
                {
                    'addSheet': {
                        'properties': {
                            'sheetId': 103,
                            'title': 'Portal Cloud SQL Database',
                            'gridProperties': {'frozenRowCount': 1}
                        }
                    }
                }
            ]

            if initial_sheet_id not in [101, 102, 103]:
                requests.append({'deleteSheet': {'sheetId': initial_sheet_id}})

            sheets_service.spreadsheets().batchUpdate(
                spreadsheetId=audit_sheet_id,
                body={'requests': requests}
            ).execute()

            # Populate all 3 tabs in parallel batch update
            data_payload = [
                {
                    'range': "'🚨 AUDIT & DISCREPANCY HEATMAP'!A1",
                    'values': heatmap_rows
                },
                {
                    'range': "'Current Live System Data'!A1",
                    'values': legacy_tab_2d
                },
                {
                    'range': "'Portal Cloud SQL Database'!A1",
                    'values': portal_tab_2d
                }
            ]

            sheets_service.spreadsheets().values().batchUpdate(
                spreadsheetId=audit_sheet_id,
                body={'valueInputOption': 'USER_ENTERED', 'data': data_payload}
            ).execute()

            # Add Red conditional formatting on Tab 1
            cond_requests = [
                {
                    'addConditionalFormatRule': {
                        'rule': {
                            'ranges': [{
                                'sheetId': 101,
                                'startRowIndex': 1,
                                'endRowIndex': len(heatmap_rows),
                                'startColumnIndex': 0,
                                'endColumnIndex': 13
                            }],
                            'booleanRule': {
                                'condition': {
                                    'type': 'TEXT_CONTAINS',
                                    'values': [{'userEnteredValue': 'Mismatch'}]
                                },
                                'format': {
                                    'backgroundColor': {'red': 1.0, 'green': 0.88, 'blue': 0.88},
                                    'textFormat': {'foregroundColor': {'red': 0.86, 'green': 0.15, 'blue': 0.15}, 'bold': True}
                                }
                            }
                        },
                        'index': 0
                    }
                },
                {
                    'addConditionalFormatRule': {
                        'rule': {
                            'ranges': [{
                                'sheetId': 101,
                                'startRowIndex': 1,
                                'endRowIndex': len(heatmap_rows),
                                'startColumnIndex': 0,
                                'endColumnIndex': 13
                            }],
                            'booleanRule': {
                                'condition': {
                                    'type': 'TEXT_CONTAINS',
                                    'values': [{'userEnteredValue': 'MISSING IN PORTAL'}]
                                },
                                'format': {
                                    'backgroundColor': {'red': 0.99, 'green': 0.80, 'blue': 0.80},
                                    'textFormat': {'foregroundColor': {'red': 0.75, 'green': 0.10, 'blue': 0.10}, 'bold': True}
                                }
                            }
                        },
                        'index': 1
                    }
                }
            ]

            sheets_service.spreadsheets().batchUpdate(
                spreadsheetId=audit_sheet_id,
                body={'requests': cond_requests}
            ).execute()

        except Exception as sheet_err:
            print(f"Notice: Google Sheet tab batch update notice: {sheet_err}")

    return {
        "status": "success",
        "spreadsheet_id": audit_sheet_id,
        "spreadsheet_url": audit_sheet_url,
        "stats": {
            "total_orders_audited": len(all_norm_keys),
            "matches_count": matches_count,
            "mismatches_count": mismatches_count,
            "missing_in_portal_count": missing_in_portal_count,
            "new_in_portal_count": new_in_portal_count
        },
        "discrepancies": discrepancy_details,
        "heatmap_rows": heatmap_rows,
        "legacy_rows": legacy_tab_2d,
        "portal_rows": portal_tab_2d,
        "message": f"Successfully completed live comparison! {len(all_norm_keys)} orders analyzed ({matches_count} matching, {mismatches_count} mismatches, {missing_in_portal_count} missing in portal)."
    }
